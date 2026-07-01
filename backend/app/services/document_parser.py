from __future__ import annotations

import csv
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

import pytesseract
from pdf2image import convert_from_path
from pypdf import PdfReader
from PIL import Image
from docx import Document as DocxDocument
from openpyxl import load_workbook


class DocumentParser:
    def __init__(self) -> None:
        self.supported_extensions = {".pdf", ".docx", ".txt", ".csv", ".xlsx", ".png", ".jpg", ".jpeg"}

    def _detect_extension(self, file_path: str | Path) -> str:
        return Path(file_path).suffix.lower()

    def _read_text_file(self, file_path: str | Path) -> str:
        return Path(file_path).read_text(encoding="utf-8", errors="ignore")

    def _read_docx(self, file_path: str | Path) -> str:
        document = DocxDocument(str(file_path))
        return "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text.strip())

    def _read_csv(self, file_path: str | Path) -> str:
        with Path(file_path).open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle))
        return "\n".join(
            ", ".join(row)
            for row in rows
            if any(cell.strip() for cell in row)
        )

    def _read_excel(self, file_path: str | Path) -> str:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheets: list[str] = []
        for sheet in workbook.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append(" | ".join("" if value is None else str(value) for value in row))
            sheets.append("\n".join(rows))
        workbook.close()
        return "\n\n".join(sheets)

    def _ocr_image(self, file_path: str | Path) -> str:
        image = Image.open(file_path)
        return pytesseract.image_to_string(image).strip()

    def _ocr_pdf(self, file_path: str | Path) -> str:
        images = convert_from_path(str(file_path))
        extracted_pages: list[str] = []
        for image in images:
            text = pytesseract.image_to_string(image).strip()
            if text:
                extracted_pages.append(text)
        return "\n\n".join(extracted_pages)

    def extract_pages(self, file_path: str | Path) -> list[dict[str, Any]]:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")

        if path.suffix.lower() not in self.supported_extensions:
            raise ValueError(f"Unsupported file type: {path.suffix or 'unknown'}")

        extension = self._detect_extension(path)
        text = ""

        if extension == ".pdf":
            try:
                reader = PdfReader(str(path))
                extracted_text_parts: list[str] = []
                for page in reader.pages:
                    page_text = (page.extract_text() or "").strip()
                    if page_text:
                        extracted_text_parts.append(page_text)
                text = "\n\n".join(extracted_text_parts)
            except Exception:
                text = ""

            if not text.strip():
                text = self._ocr_pdf(path)

        elif extension == ".docx":
            text = self._read_docx(path)
        elif extension == ".txt":
            text = self._read_text_file(path)
        elif extension == ".csv":
            text = self._read_csv(path)
        elif extension == ".xlsx":
            text = self._read_excel(path)
        elif extension in {".png", ".jpg", ".jpeg"}:
            text = self._ocr_image(path)

        if not text.strip():
            raise ValueError("No extractable text was found in this file.")

        return [{"page_number": 1, "text": text}]

    def extract_text(self, file_path: str | Path) -> str:
        pages = self.extract_pages(file_path)
        return "\n\n".join(page["text"] for page in pages if page.get("text"))
